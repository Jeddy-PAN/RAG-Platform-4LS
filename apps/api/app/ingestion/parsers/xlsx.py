from pathlib import Path

from app.ingestion.parsers.base import ParserError, NormalizedSection


class XlsxParser:
    """Parser for visible XLSX worksheet rows."""

    def parse(self, path: Path) -> list[NormalizedSection]:
        from openpyxl import load_workbook

        workbook = load_workbook(path, data_only=True, read_only=True)
        sections: list[NormalizedSection] = []
        for sheet in workbook.worksheets:
            if sheet.sheet_state != "visible":
                continue
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                if not values:
                    continue
                sections.append(
                    NormalizedSection(
                        section_index=len(sections),
                        text=" | ".join(values),
                        source_metadata={
                            "sheet_name": sheet.title,
                            "row_start": row_index,
                            "row_end": row_index,
                        },
                    )
                )
        workbook.close()

        if not sections:
            raise ParserError("Document contains no usable text")
        return sections
